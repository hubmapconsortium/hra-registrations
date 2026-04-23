#!/usr/bin/env python3
"""
Generate a KPMP `registrations.yaml` from the HuBMAP Healthy Tissue Registry
spreadsheet, conforming to:
  https://raw.githubusercontent.com/hubmapconsortium/hra-rui-locations-processor/main/registrations.schema.json

Normalizes dataset technology names against the HRA canonical list:
  https://apps.humanatlas.io/api/v1/technology-names

Deduplicates `(canonical_technology, link)` pairs per donor.

Usage:
  python hrt_to_registrations.py <path/to/HuBMAP_Healthy_Tissue_Registry.xlsx>

Output:
  registrations.yaml written next to the input spreadsheet.
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required. Install with: pip install openpyxl")


# ──────────────────────────────────────────────────────────────────────────
# Provider metadata — KPMP-specific defaults carried over from the prior
# registrations.yaml. Change these if re-purposing this script.
# ──────────────────────────────────────────────────────────────────────────
PROVIDER = {
    "consortium_name": "KPMP",
    "provider_name": "KPMP",
    "provider_uuid": "d026da42-1604-41f0-ace4-045b051fe96b",
    "defaults": {
        "id": "cc27b52b-d98e-4fa7-97d7-142ab2665f6b",
        "thumbnail": "https://cdn.humanatlas.io/ui/ccf-eui/assets/icons/ico-unknown.svg",
        "link": "https://atlas.kpmp.org/",
    },
}

# Map from xlsx "Dataset Technology" values to canonical HRA names.
# See: https://apps.humanatlas.io/api/v1/technology-names
TECH_MAP = {
    # RNAseq family
    "Single-cell RNA-Seq": "RNAseq",
    "Single-nucleus RNA-Seq": "RNAseq",
    "Atlas v2 Single-cell RNAseq": "RNAseq",
    "Atlas v2 Single-nucleaus RNAseq": "RNAseq",  # note xlsx typo preserved as KEY
    "CZ CELLxGENE Atlas v2 Single-cell RNAseq": "RNAseq",
    "CZ CELLxGENE Atlas v2 Single-nucleus RNAseq": "RNAseq",
    # ATAC / Multiome
    "ATAC-Seq": "ATAC-seq",
    "10x Multiome": "10x",
    # Imaging
    "CODEX": "CODEX",
    "Imaging Mass Cytometry": "Imaging Mass Cytometry",
    "Light Microscopic Whole Slide Images": "Histology",
    "3D Tissue Imaging and Cytometry": "Light Sheet",
    "3D Tissue Imaging and Cytometry No Channels": "Light Sheet",
    # Spatial-omics
    "Spatial Transcriptomics": "Spatial Transcriptomics",
    "Multimodal Imaging Mass Spectrometry": "MIMS",
    "Spatial Lipidomics": "MALDI-IMS",
    "Spatial Metabolomics": "MALDI-IMS",
    "Spatial N-Glycomics": "MALDI-IMS",
    # Not tied to a spatial assay — bucketed as OTHER
    "Clinical Chemistry": "OTHER",
    "MSD Plasma Biomarker": "OTHER",
    "MSD Urine Biomarker": "OTHER",
    "Metabolon Plasma": "OTHER",
    "SomaScan Plasma Proteomics": "OTHER",
    "SomaScan Urine Proteomics": "OTHER",
    "Pathology Descriptor Scoring": "OTHER",
    "Segmentation Masks & Pathomics Vectors": "OTHER",
}

# xlsx "Sample Location" → token for the rui_location JSON filename.
# ("Sample Location" holds Upper pole/Middle/Lower pole;
#  "Sample Laterality" holds Left/Right.)
AXIS_TOKEN = {"Upper pole": "upper", "Middle": "mid", "Lower pole": "lower"}


def rui_filename(sex: str, axis: str, side: str) -> str:
    """Produce filename matching the existing HRA registration convention:
       male-upper-left-pole.json, female-mid-right.json, etc."""
    tx = sex.lower()
    ts = side.lower()
    ta = AXIS_TOKEN[axis]
    return f"{tx}-mid-{ts}.json" if ta == "mid" else f"{tx}-{ta}-{ts}-pole.json"


def nat_key(donor_id: str) -> tuple[int, int]:
    """Natural sort key for IDs like '163-7', '164-10', '165-15'."""
    a, b = donor_id.split("-")
    return int(a), int(b)


def read_combined_sheet(xlsx_path: Path) -> list[dict]:
    """Read the 'Combined data' sheet and return a list of row dicts."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Combined data" not in wb.sheetnames:
        sys.exit(
            f"Expected sheet 'Combined data' in {xlsx_path}. "
            f"Found: {wb.sheetnames}"
        )
    ws = wb["Combined data"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit(f"Sheet 'Combined data' is empty in {xlsx_path}")
    header = list(rows[0])
    return [dict(zip(header, r)) for r in rows[1:] if r and r[0]]


def build_donors(rows: list[dict]) -> tuple[dict, int]:
    """Group rows by donor; normalize tech; dedupe (canonical_tech, link) per donor.
    Returns (donor_map, n_duplicates_removed)."""
    donors: dict = defaultdict(
        lambda: {
            "sex": None,
            "age": None,
            "axis": None,
            "side": None,
            "datasets": [],
            "_seen": set(),
        }
    )
    duplicates = 0
    unknown_techs: set[str] = set()

    for r in rows:
        did = r["Donor ID"]
        d = donors[did]
        d["sex"] = r["Donor Sex"]
        d["age"] = r["Donor Age"]
        d["axis"] = r["Sample Location"]    # Upper pole / Middle / Lower pole
        d["side"] = r["Sample Laterality"]  # Left / Right

        raw_tech = r["Dataset Technology"]
        link = r["Dataset Deep Link"]
        canonical = TECH_MAP.get(raw_tech)
        if canonical is None:
            unknown_techs.add(raw_tech)
            canonical = "OTHER"

        key = (canonical, link)
        if key in d["_seen"]:
            duplicates += 1
            continue
        d["_seen"].add(key)
        d["datasets"].append({"technology": canonical, "link": link})

    if unknown_techs:
        print(
            f"WARNING: {len(unknown_techs)} technology names not in TECH_MAP "
            f"(defaulted to OTHER): {sorted(unknown_techs)}",
            file=sys.stderr,
        )
    return donors, duplicates


def emit_yaml(donors: dict, out_path: Path) -> tuple[int, int]:
    """Write the registrations.yaml file. Returns (n_donors, n_datasets)."""
    # Hand-rolled YAML to match the key ordering / indentation of the existing
    # registrations.yaml files in the HRA ecosystem. Values here never contain
    # YAML-significant leading characters or line breaks, so bare scalars are
    # safe.
    lines = [
        "# yaml-language-server: $schema=https://raw.githubusercontent.com/hubmapconsortium/hra-rui-locations-processor/main/registrations.schema.json",
        "",
        f"- consortium_name: {PROVIDER['consortium_name']}",
        f"  provider_name: {PROVIDER['provider_name']}",
        f"  provider_uuid: {PROVIDER['provider_uuid']}",
        "  defaults:",
        f"    id: {PROVIDER['defaults']['id']}",
        f"    thumbnail: {PROVIDER['defaults']['thumbnail']}",
        f"    link: {PROVIDER['defaults']['link']}",
        "  donors:",
    ]

    n_datasets = 0
    for donor_idx, did in enumerate(sorted(donors.keys(), key=nat_key), start=1):
        d = donors[did]
        fname = rui_filename(d["sex"], d["axis"], d["side"])
        donor_id = f"https://atlas.kpmp.org/Donor#{donor_idx}"
        sample_id = f"{donor_id}_Block#1"
        lines += [
            f"  - id: {donor_id}",
            f"    label: HRT {d['age']}",
            f"    link: https://atlas.kpmp.org/#{did}",
            f"    sex: {d['sex']}",
            "    samples:",
            f"    - id: {sample_id}",
            f"      rui_location: {fname}",
            "      datasets:",
        ]
        # Stable order: by technology, then by link
        ds_sorted = sorted(d["datasets"], key=lambda x: (x["technology"], x["link"]))
        for ds_idx, ds in enumerate(ds_sorted, start=1):
            n_datasets += 1
            dataset_id = f"{sample_id}_Dataset#{ds_idx}"
            lines += [
                f"      - id: {dataset_id}",
                f"        link: {ds['link']}",
                f"        technology: {ds['technology']}",
                f"        label: {ds['technology']}",
            ]

    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return len(donors), n_datasets


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a KPMP registrations.yaml from the HuBMAP Healthy "
                    "Tissue Registry xlsx."
    )
    parser.add_argument(
        "xlsx",
        type=Path,
        help="Path to HuBMAP_Healthy_Tissue_Registry.xlsx (must contain a "
             "'Combined data' sheet).",
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        default=None,
        help="Output path for registrations.yaml. Defaults to "
             "<xlsx-dir>/registrations.yaml.",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"Input file not found: {args.xlsx}")

    out_path = args.output or args.xlsx.parent / "registrations.yaml"

    rows = read_combined_sheet(args.xlsx)
    donors, duplicates = build_donors(rows)
    n_donors, n_datasets = emit_yaml(donors, out_path)

    print(f"Input rows:         {len(rows)}")
    print(f"Duplicates removed: {duplicates}")
    print(f"Donors:             {n_donors}")
    print(f"Datasets:           {n_datasets}")
    print(f"Wrote:              {out_path}")


if __name__ == "__main__":
    main()
